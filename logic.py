from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime
import shutil
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
#from os import mkdir,path,walk
def copy_file(source, destination):
    shutil.copy(source, destination)
def get_date(date=None):
    if date:
        today = datetime.datetime.strptime(str(date), '%Y-%m-%d')
    else:
        today = datetime.datetime.today()
    end_time = today - datetime.timedelta(days=today.isoweekday())
    start_time = end_time - datetime.timedelta(days=6)
    return start_time.strftime("%Y-%m-%d"), end_time.strftime("%Y-%m-%d")

print(get_date(''))
def GDPS(readpath,readexeclpath,writepath):
    num=2
    #copy_file(readexeclpath,writepath)  
    with open(readpath, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)
        table = excel.get_sheet_by_name('GDPS')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(vertical="center")
        while line:
            if line=='':
                break
            linelist=line.split()[:-1]
            #print(linelist)
            table['A'+str(num)]=linelist[0]
            table['A'+str(num)].number_format='yyyy/mm/dd'
            table['B'+str(num)]=linelist[1]
            table['B'+str(num)].number_format='hh:mm:ss'
            table['C'+str(num)]=linelist[2]
            table['C'+str(num)].number_format='@'
            table['A'+str(num)].font=vt_font
            table['A'+str(num)].alignment = alignment1
            table['B'+str(num)].font=vt_font
            table['B'+str(num)].alignment = alignment2
            table['C'+str(num)].font=vt_font
            table['C'+str(num)].alignment = alignment2
            if(len(linelist)>3):
                table['D'+str(num)]=linelist[3]
                table['D'+str(num)].number_format='@'    
                table['D'+str(num)].font=vt_font  
                table['D'+str(num)].alignment = alignment2 
            if num==2:
                date1=linelist[0]
                print(date1)    
            num+=1
            line=file.readline()
        date2=table['A'+str(num-1)].value
        print(date2)
        excel.save(writepath+''.join(date1.split('/')[1:])+'_'+''.join(date2.split('/')[1:])+'_'+'GDPS.xlsx')

def MCDS(readpath,readexeclpath,writepath):
    num=1
    #copy_file(readexeclpath,writepath)  
    with open(readpath, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)
        table = excel.get_sheet_by_name('MCDS')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(vertical="center")
        while line:
            if line=='':
                break
            #print(linelist)
            table['A'+str(num)]=line
            table['A'+str(num)].font=vt_font
            table['A'+str(num)].alignment = alignment2
            num+=1
            line=file.readline()
        today = datetime.datetime.today()
        time = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=6)
        date1=''.join(str(time).split()[0].split('-')[1:])
        excel.save(writepath+date1+'_'+'MCDS.xlsx')

def CGFAIL(readpath,readexeclpath,writepath):
    num=2
    #copy_file(readexeclpath,writepath)  
    with open(readpath, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)
        table = excel.get_sheet_by_name('CGFAIL')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(vertical="center")
        while line:
            if line=='':
                break
            linelist=line.split()
            #print(linelist)
            #datetime.datetime.strptime(linelist[0],'%Y/%m/%d')
            table['A'+str(num)]='20'+linelist[0]
            table['A'+str(num)].number_format='yyyy/mm/dd'
            table['B'+str(num)]=linelist[1]
            table['B'+str(num)].number_format='hh:mm:ss'
            table['C'+str(num)]=int(linelist[2])
            table['A'+str(num)].font=vt_font
            table['A'+str(num)].alignment = alignment1
            table['B'+str(num)].font=vt_font
            table['B'+str(num)].alignment = alignment1
            table['C'+str(num)].font=vt_font
            table['C'+str(num)].alignment = alignment1 
            num+=1
            line=file.readline()
        today = datetime.datetime.today()
        time1 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=7)
        time2 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=1)
        date_str=''.join(str(time1).split()[0].split('-')[1:])+'_'+''.join(str(time2).split()[0].split('-')[1:])
        excel.save(writepath+date_str+'_'+'CGFAIL.xlsx')

def DSPL(readpath,readexeclpath,writepath):
    num=2
    #copy_file(readexeclpath,writepath)  
    with open(readpath, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)
        table = excel.get_sheet_by_name('DSPL')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(vertical="center")
        while line:
            if line=='':
                break
            linelist=line.split()
            #print(linelist)
            #datetime.datetime.strptime(linelist[0],'%Y/%m/%d')
            table['A'+str(num)]='20'+linelist[0]
            table['B'+str(num)]=linelist[1]+':00'
            table['A'+str(num)].font=vt_font
            table['A'+str(num)].alignment = alignment1
            table['B'+str(num)].font=vt_font
            table['B'+str(num)].alignment = alignment1
            for i in range(2,len(linelist)):
                table[chr(ord('A')+i)+str(num)]=round(float(linelist[i]), 2)
                table[chr(ord('A')+i)+str(num)].font=vt_font
                table[chr(ord('A')+i)+str(num)].alignment = alignment1
            num+=1
            line=file.readline()
        today = datetime.datetime.today()
        time1 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=7)
        time2 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=1)
        date_str=''.join(str(time1).split()[0].split('-')[1:])+'_'+''.join(str(time2).split()[0].split('-')[1:])
        excel.save(writepath+date_str+'_'+'DSPL.xlsx')

def SMSSGCHK(readpath,readexeclpath,writepath):
    num=2
    #copy_file(readexeclpath,writepath)  
    with open(readpath, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)
        table = excel.get_sheet_by_name('SMSSGCHK')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(horizontal="right",vertical="center")
        alignment3 = Alignment(vertical="center")
        while line:
            if line=='':
                break
            linelist=line.split()

            table['A'+str(num)]='20'+linelist[0]
            table['A'+str(num)].number_format='yyyy/m/d'
            table['B'+str(num)]=linelist[1].split('.')[0]
            table['A'+str(num)].font=vt_font
            table['A'+str(num)].alignment = alignment1
            table['B'+str(num)].font=vt_font
            table['B'+str(num)].alignment = alignment1
            for i in range(2,len(linelist)):
                if '%' in linelist[i]:
                    linelist[i]=int(linelist[i].replace('%',''))/100
                table[chr(ord('A')+i)+str(num)]=linelist[i]
                table[chr(ord('A')+i)+str(num)].font=vt_font
                table[chr(ord('A')+i)+str(num)].alignment = alignment1
            table['E'+str(num)].alignment = alignment2
            table['F'+str(num)].alignment = alignment3
            table['G'+str(num)].alignment = alignment3
            table['F'+str(num)].number_format='0%'
            table['G'+str(num)].number_format='0%'
            num+=1
            line=file.readline()

        today = datetime.datetime.today()
        time1 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=7)
        time2 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=1)
        date_str=''.join(str(time1).split()[0].split('-')[1:])+'_'+''.join(str(time2).split()[0].split('-')[1:])
        excel.save(writepath+date_str+'_'+'SMSSGCHK.xlsx')

def STORAGE(readpath,readexeclpath,writepath):
    num=2
    readpath1='STORAGE1.txt'
    readpath2='STORAGE2.txt'  
    datadict={}
    with open(readpath+'/'+readpath1, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)

        table = excel.get_sheet_by_name('STORAGE（LPAR順）')
        table1 = excel.get_sheet_by_name('STORAGE（日付順）')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        fill1 = PatternFill(patternType='solid', fgColor='E2EFDA')
        fill2 = PatternFill(patternType='solid', fgColor='FCE4D6')
        fill3 = PatternFill(patternType='solid', fgColor='D9E1F2')
        side1 = Side(style='thin', color='000000')
        side2=Side(style='medium',color='000000')
        border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
        border2= Border(top=side2, bottom=side1, left=side1, right=side1)
        border3= Border(top=side2, bottom=side1, left=side1, right=side2)
        border4= Border(top=side1, bottom=side1, left=side1, right=side2)
        border5= Border(top=side1, bottom=side2, left=side1, right=side1)
        border6= Border(top=side1, bottom=side2, left=side1, right=side2)
        border7= Border(top=side1, bottom=side2, left=side2, right=side1)
        border8= Border(top=side1, bottom=side1, left=side2, right=side1)
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(horizontal="right",vertical="center")
        record=''
        while line:
            if line=='':
                for i in range(8):
                    if i in [2,3]:
                        table[chr(ord('A')+i)+str(num)].fill=fill1
                    if i in [4,5]:
                        table[chr(ord('A')+i)+str(num)].fill=fill2
                    if i in [6,7]:
                        table[chr(ord('A')+i)+str(num)].fill=fill3
                    table[chr(ord('A')+i)+str(num)].border=border2 
                table[chr(ord('A')+8)+str(num)].border=border3    
                break
            linelist=line.split()
            if linelist[1] in datadict.keys():
                datadict[linelist[1]].append([linelist[0]]+linelist[2:])
            else:
                datadict[linelist[1]]=[[linelist[0]]+linelist[2:]]

            table['A'+str(num)]=linelist[0]

            table['B'+str(num)].value='20'+linelist[1]
            
            table['A'+str(num)].font=vt_font
            table['A'+str(num)].alignment = alignment1

            table['B'+str(num)].font=vt_font
            table['B'+str(num)].alignment = alignment1
            table['B'+str(num)].number_format='yyyy/mm/dd'
            table['B'+str(num)].border=border1
            
            for i in range(2,len(linelist)):
                if i in [3,5,7,8]:
                    table[chr(ord('A')+i)+str(num)]=int(linelist[i])
                else:
                    table[chr(ord('A')+i)+str(num)]=linelist[i]
                    #table1[chr(ord('A')+i)+str(num)]=linelist[i]
                table[chr(ord('A')+i)+str(num)].font=vt_font
                table[chr(ord('A')+i)+str(num)].alignment = alignment1
                table[chr(ord('A')+i)+str(num)].border=border1
                if i in [2,3]:
                    table[chr(ord('A')+i)+str(num)].fill=fill1
                if i in [4,5]:
                    table[chr(ord('A')+i)+str(num)].fill=fill2
                if i in [6,7]:
                    table[chr(ord('A')+i)+str(num)].fill=fill3
                if i in [2,4,6]:
                    table[chr(ord('A')+i)+str(num)].number_format='hh:mm'
                if i in [3,5,7]:
                    table[chr(ord('A')+i)+str(num)].number_format='0_);[red](0)'
                if i==8:
                    table[chr(ord('A')+i)+str(num)].number_format='0_ '
            table[chr(ord('A')+8)+str(num)].border=border4
            table[chr(ord('A')+8)+str(num)].alignment = alignment2
            if record!=linelist[0]:
                for i in range(8):
                   table[chr(ord('A')+i)+str(num)].border=border2 
                table[chr(ord('A')+8)+str(num)].border=border3
                record=linelist[0]

            num+=1
            line=file.readline() 
        sort_datadict=dict(sorted(datadict.items()))
        num=2
        for key in sort_datadict.keys():
            for item in sort_datadict[key]:
                table1['A'+str(num)]='20'+key
                table1['B'+str(num)].value=item[0]
                
                table1['A'+str(num)].font=vt_font
                table1['A'+str(num)].alignment = alignment1

                table1['B'+str(num)].font=vt_font
                table1['B'+str(num)].alignment = alignment1
                table1['A'+str(num)].number_format='yyyy/mm/dd'
                table1['A'+str(num)].border=border1
                table1['B'+str(num)].border=border8
                for i in range(1,len(item)):
                    if i in [2,4,6,7]:
                        table1[chr(ord('A')+i+1)+str(num)]=int(item[i])
                    else:
                        table1[chr(ord('A')+i+1)+str(num)]=item[i]
                    table1[chr(ord('A')+i+1)+str(num)].font=vt_font
                    table1[chr(ord('A')+i+1)+str(num)].alignment = alignment1
                    table1[chr(ord('A')+i+1)+str(num)].border=border1
                    if i in [1,2]:
                        table1[chr(ord('A')+i+1)+str(num)].fill=fill1
                    if i in [3,4]:
                        table1[chr(ord('A')+i+1)+str(num)].fill=fill2
                    if i in [5,6]:
                        table1[chr(ord('A')+i+1)+str(num)].fill=fill3
                    if i in [1,3,5]:
                        table1[chr(ord('A')+i+1)+str(num)].number_format='hh:mm'
                    if i in [2,4,6]:
                        table1[chr(ord('A')+i+1)+str(num)].number_format='0_);[red](0)'
                    if i==7:                       
                        table1[chr(ord('A')+i+1)+str(num)].number_format='0_ '
                        table1[chr(ord('A')+i+1)+str(num)].border=border4
                num+=1
            for i in range(8):
                if i==1:
                    table1[chr(ord('A')+i)+str(num-1)].border=border7
                else:
                    table1[chr(ord('A')+i)+str(num-1)].border=border5 
            table1[chr(ord('A')+8)+str(num-1)].border=border6
            # print(sort_datadict)

    with open(readpath+'/'+readpath2, 'r') as file:
        line=file.readline()
        num=2
        #excel=load_workbook(readexeclpath)
        table2 = excel.get_sheet_by_name('MSGID CHECK')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(vertical="center")
        while line:
            if line=='':
                break
            linelist=line.split()

            table2['A'+str(num)]=linelist[0]
            table2['B'+str(num)]=linelist[1]
            table2['A'+str(num)].font=vt_font
            table2['A'+str(num)].alignment = alignment1
            table2['B'+str(num)].font=vt_font
            table2['B'+str(num)].alignment = alignment1
            for i in range(2,len(linelist)):
                if i==4:
                    table2[chr(ord('A')+i)+str(num)]=' '.join(linelist[4:])
                    table2[chr(ord('A')+i)+str(num)].font=vt_font
                    table2[chr(ord('A')+i)+str(num)].alignment = alignment1
                    break  
                else:
                    table2[chr(ord('A')+i)+str(num)]=linelist[i]
                table2[chr(ord('A')+i)+str(num)].font=vt_font
                table2[chr(ord('A')+i)+str(num)].alignment = alignment1
                if i==2:
                    table2[chr(ord('A')+i)+str(num)].number_format='yyyy/mm/dd'
                if i==3:
                    table2[chr(ord('A')+i)+str(num)].number_format='hh:mm'
            num+=1
            line=file.readline()


        today = datetime.datetime.today()
        time1 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=7)
        time2 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=1)
        date_str=''.join(str(time1).split()[0].split('-')[1:])+'_'+''.join(str(time2).split()[0].split('-')[1:])
        excel.save(writepath+date_str+'_'+'STORAGE.xlsx')
def search_index(list,value,col,givenIndex=0):
    for i in range(givenIndex,len(list)):
        
        if col==0 and value==str(list[i][col])[5:]:
            #print(value,list[i][col])
            return i
        if col==1 and str(value)==str(list[i][col]):
            #print(value,list[i][col])
            return i
def RMFSUM(readpath,readexeclpath,readexcelpath1,writepath):
    num=2
    #copy_file(readexeclpath,writepath) 
    side1 = Side(style='thin', color='000000') 
    border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
    with open(readpath, 'r') as file:
        line=file.readline()
        excel=load_workbook(readexeclpath)
        table = excel.get_sheet_by_name('RMFSUM')
        vt_font = Font(name="ＶＴ ゴシック",size=11,bold=False,italic=False,color="000000")
        alignment1 = Alignment(horizontal="center",vertical="center")
        alignment2 = Alignment(vertical="center")
        line_list=[]
        value_list=[]
        index_list=[]
        while line:
            if line=='':
                break
            #linelist=line.split()
            #print(linelist)
            #datetime.datetime.strptime(linelist[0],'%Y/%m/%d')

            line_list.append(line)
            num+=1
            line=file.readline()
        middle_str=line_list[len(line_list)//2]
        for i in range(len(middle_str)):
            if i!=0 and middle_str[i-1]!=' ' and middle_str[i]==' ':
                index_list.append(i)
        #print(index_list)
        for i in line_list:
            temp_list=[]
            for j in range(len(index_list)+1):
                if j==len(index_list):
                    temp_value=i[index_list[j-1]:]
                elif j>0:
                    temp_value=i[index_list[j-1]:index_list[j]]
                # elif j==len(index_list):
                #     temp_value=i[index_list[j]:]
                else:
                    temp_value=i[:index_list[j]]
                
                temp_value=temp_value.replace(' ','')
                temp_value=temp_value.replace('\n','')
                if j>0 and temp_value!='':
                    temp_value=float(temp_value)
                temp_list.append(temp_value)
            
            value_list.append(temp_list)
        #print(value_list[0])
        #print(line_list[0])
    lastweek_excel=load_workbook(readexcelpath1)

    table1 = lastweek_excel.get_sheet_by_name('RMFSUM')
    num=3
    #num=10459
    value_list1=[]
    flag=True
    # print(table1['A'+str(num)].value)
    while flag:
        line_list1=[]
        for i in range(29):
            if i==0 and table1['A'+str(num)].value=='_x001A_':
                flag=False
                break
            elif i==0:
                line_list1.append(table1.cell(row=num,column=i+1).value.strftime('%Y/%m/%d'))
            else:
                line_list1.append(table1.cell(row=num,column=i+1).value)
            
        if flag:
            value_list1.append(line_list1)
            num+=1
    
    index_row=search_index(value_list1,value_list[0][0],col=0)
    #print(type(index_row))
    #print(index_row)
    index_row=search_index(value_list1,value_list[0][1],col=1,givenIndex=index_row)
    #print(index_row)
    for i,line in enumerate(value_list):
        if '' not in line:
            break
        for j,vau in enumerate(line):
            if vau=='':
                value_list[i][j]=value_list1[i+index_row][j]
    
    #print(value_list[0])
    nowTimeStr=datetime.datetime.today().strftime('%Y/%m/%d')
    for i,line in enumerate(value_list):

        for j,vau in enumerate(line):
            table.cell(row=i+3,column=j+1).value=vau
            table.cell(row=i+3,column=j+1).border=border1
            table.cell(row=i+3,column=j+1).font=vt_font
            table.cell(row=i+3,column=j+1).alignment = alignment2

            if j==0:
                timestr=nowTimeStr[:4]+'/'+table.cell(row=i+3,column=j+1).value
                dt = datetime.datetime.strptime(timestr, "%Y/%m/%d") 
                table.cell(row=i+3,column=j+1).value=dt
                table.cell(row=i+3,column=j+1).number_format='yyyy/mm/dd'
            elif j==1:
                table.cell(row=i+3,column=j+1).alignment = alignment1
                table.cell(row=i+3,column=j+1).number_format='00.00'
            else:
                table.cell(row=i+3,column=j+1).number_format='0.0_);[red](0.0)'



   
    today = datetime.datetime.today()
    time1 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=7)
    time2 = today - datetime.timedelta(days=today.isoweekday())-datetime.timedelta(days=1)
    date_str=''.join(str(time1).split()[0].split('-')[1:])+'_'+''.join(str(time2).split()[0].split('-')[1:])
    excel.save(writepath+date_str+'_'+'RMFSUM.xlsx')

now_time=datetime.datetime.today()
# print(str(now_time)[2:10])
# print(datetime.datetime.today())
# GDPS('项目组相关/CAPA/モニタリング/GDPS/GDPS.txt','项目组相关/CAPA/excel_input/mmdd1_mmdd2_GDPS.xlsx','项目组相关/CAPA/excel_output/testGDPS.xlsx')
# MCDS('项目组相关/CAPA/モニタリング/MCDS/MCDS.txt','项目组相关/CAPA/excel_input/mmdd_MCDS.xlsx','项目组相关/CAPA/excel_output/testMCDS.xlsx')
# CGFAIL('项目组相关/CAPA/モニタリング/CGFAIL/CGFAIL.txt','项目组相关/CAPA/excel_input/mmdd1_mmdd2_CGFAIL.xlsx','项目组相关/CAPA/excel_output/testCGFAIL.xlsx')
# DSPL('项目组相关/CAPA/モニタリング/DSPL/DSPL.txt','项目组相关/CAPA/excel_input/mmdd1_mmdd2_DSPL.xlsx','项目组相关/CAPA/excel_output/testDSPL.xlsx')
# SMSSGCHK('项目组相关/CAPA/モニタリング/SMSSGCHK/SMSSGCHK.txt','项目组相关/CAPA/excel_input/mmdd1_mmdd2_SMSSGCHK.xlsx','项目组相关/CAPA/excel_output/testSMSSGCHK.xlsx')       
# STORAGE('项目组相关/CAPA/モニタリング/STORAGE','项目组相关/CAPA/excel_input/mmdd1_mmdd2_STORAGE.xlsx','项目组相关/CAPA/excel_output/testSTORAGE.xlsx')
# RMFSUM('C:/vscode_workspace/项目组相关/CAPA/モニタリング/RMFSUM/RMFSUM.txt','C:/vscode_workspace/项目组相关/CAPA/excel_input/mmdd1_mmdd2_RMFSUM.xlsx','C:\ホスト基盤\キャパシティー管理\モニタリング240304\モニタリング\\0225_0302_RMFSUM.xlsx')
# print(datetime.datetime.today())
# print('please input the input folder root')
# #input_file_root=input()
# input_file_root='C:/vscode_workspace\项目组相关\CAPA\モニタリング'
# generate_boolean_dict={'GDPS':True,'MCDS':True,'CGFAIL':True,'DSPL':True,'SMSSGCHK':True,'STORAGE':True,'RMFSUM':True}
# print('please input the output folder root')
# #output_file_root=input()
# output_file_root='C:/vscode_workspace\项目组相关\CAPA/txt_input'
# output_folder_path=output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング'
# if not path.exists(output_folder_path) or not path.isdir(output_folder_path):
#     mkdir(output_folder_path)
# if generate_boolean_dict['GDPS']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'GDPS' in filename:
#                 print('GDPS is exist')
#                 flag=False
#     if flag:
#         GDPS(input_file_root+'/GDPS/GDPS.txt','excel_input/mmdd1_mmdd2_GDPS.xlsx',output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')
        
# if generate_boolean_dict['MCDS']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'MCDS' in filename:
#                 print('MCDS is exist')
#                 flag=False
#     if flag:
#         MCDS(input_file_root+'/MCDS/MCDS.txt','excel_input/mmdd_MCDS.xlsx',output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')

# if generate_boolean_dict['CGFAIL']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'CGFAIL' in filename:
#                 print('CGFAIL is exist')
#                 flag=False
#     if flag:
#         CGFAIL(input_file_root+'/CGFAIL/CGFAIL.txt','excel_input/mmdd1_mmdd2_CGFAIL.xlsx',output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')

# if generate_boolean_dict['DSPL']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'DSPL' in filename:
#                 print('DSPL is exist')
#                 flag=False
#     if flag:
#         DSPL(input_file_root+'/DSPL/DSPL.txt','excel_input/mmdd1_mmdd2_DSPL.xlsx',output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')

# if generate_boolean_dict['SMSSGCHK']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'SMSSGCHK' in filename:
#                 print('SMSSGCHK is exist')
#                 flag=False
#     if flag:
#         SMSSGCHK(input_file_root+'/SMSSGCHK/SMSSGCHK.txt','excel_input/mmdd1_mmdd2_SMSSGCHK.xlsx',output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')

# if generate_boolean_dict['STORAGE']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'STORAGE' in filename:
#                 print('STORAGE is exist')
#                 flag=False
#     if flag:
#         STORAGE(input_file_root+'/STORAGE','excel_input/mmdd1_mmdd2_STORAGE.xlsx',output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')

# if generate_boolean_dict['RMFSUM']:
#     flag=True
#     for _,_,filelist in walk(output_folder_path):
#         for filename in filelist:
#             if 'RMFSUM' in filename:
#                 print('RMFSUM is exist')
#                 flag=False
#     rmfsum_last_week=input('please input RMFSUM file last week')
#     if flag:
#         RMFSUM(input_file_root+'/RMFSUM/RMFSUM.txt','excel_input/mmdd1_mmdd2_RMFSUM.xlsx',rmfsum_last_week,output_file_root+'/'+str(now_time)[2:10].replace('-','')+'モニタリング/')
# now_time1=datetime.datetime.today()
# delta=now_time1-now_time

# print('这次生成一共花费了'+str((delta.seconds//60) % 60)+'分',str(delta.seconds%60)+'秒')

