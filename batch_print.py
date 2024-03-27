import xlwings as xw
import openpyxl
import os
def getsheetname(filename):
    #workbook = openpyxl.load_workbook('C:/ホスト基盤/test/0001.xlsx')	# 返回一个workbook数据类型的值
    workbook = openpyxl.load_workbook(filename)	
    print_sheets=[]
    for i in workbook.sheetnames:

        if '印刷対象外' in i:
            break
        else:
            #print(i)
            print_sheets.append(i)

    #print(print_sheets)
    workbook.close()
    return print_sheets

def printfile(filename):
    print_sheets=getsheetname(filename)
    app = xw.App(visible=False, add_book=False)
    workbook = app.books.open(filename)

    worksheet = workbook.sheets(print_sheets)
    
    worksheet.api.PrintOut(Copies=1, ActivePrinter='', Collate=True)
    workbook.close()
    app.quit()

def printfolder(folder):
    msg=[]
    excel_files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
    for file in excel_files:
        filename = os.path.join(folder, file)
        filename = folder+ '/'+file
        print(filename)
        try:
            printfile(filename)
            msg.append(filename.split('/')[-1]+'印刷完了')
        except Exception as e:
            msg.append(filename.split('/')[-1]+'印刷失敗')

    return msg
#printfolder('C:/ホスト基盤/test')