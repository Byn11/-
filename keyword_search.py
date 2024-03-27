from os import mkdir,listdir,makedirs
from shutil import copy
from openpyxl import load_workbook
msg=[]
def copy_file(source, destination):
    copy(source, destination)

def getname(in_path):
    keys,values=[],[]
    with open(in_path, 'r',encoding='UTF-8') as file:
        line=file.readline()
        while line:
            keys.append(line.split('_')[0])
            values.append(line.split('_')[1].strip('\n'))
            line=file.readline()
        my_dict = dict(zip(keys, values))
    return my_dict


dir_path = 'keyword_search/product_information.txt'
#print(getname(dir_path))

def makefile(destination):
    copy_file('keyword_search/SAMPLE.xlsx',destination)
def makefileGNP(destination):
    copy_file('keyword_search/GNPSAMPLE.xlsx',destination)

def rwfile(readpath,folder,isGNP):
    print('***********************************888')
    print(folder)
    with open(readpath, 'r') as file:
        num=9
        line=file.readline()
        pre_keyword=''.join(line.split()[2:-1])
        now_keyword=pre_keyword
        writepath=folder+'/'+folder.split('/')[-1]+'_'+now_keyword.replace("\\","¥").replace("/","").replace("*","")+'.xlsx'
        if isGNP:
            #print('***********************************888')
            print(writepath)
            makefileGNP(writepath)
        else:
            
            print(writepath)
            makefile(writepath)
        excel=load_workbook(writepath)
        table = excel.get_sheet_by_name('検索結果')
        table['D4']=now_keyword
        str1=''.join(folder.split('/')[-1].split('_')[1:])
        str1=' '.join(str1.split()[:-1])
        table['A1']='影響調査結果（'+str1+'）'
        while line:
            if line=='':
                break
            #print(line)
            now_keyword=''.join(line.split()[2:-1])
            if pre_keyword!=now_keyword:
                excel.save(writepath)
                writepath=folder+'/'+folder.split('/')[-1]+'_'+now_keyword.replace("\\","¥").replace("/","").replace("*","")+'.xlsx'
                if isGNP:
                    makefileGNP(writepath)
                else:
                    makefile(writepath)
                num=9
                excel=load_workbook(writepath)
                table = excel.get_sheet_by_name('検索結果')
                table['D4']=now_keyword
                str1=''.join(folder.split('/')[-1].split('_')[1:])
                str1=' '.join(str1.split()[:-1])
                table['A1']='影響調査結果（'+str1+'）'
            #print(table)
            table['D'+str(num)]=line.split()[0]
            table['E'+str(num)]=line.split()[-1]
            table['F'+str(num)]=int(line.split()[1])

            num+=1
            pre_keyword=now_keyword
            line=file.readline()

        excel.save(writepath)
        msg.append(writepath.split('/')[-1]+'生成完了')
        # connection.msgtran(msg)

#rwfile('検索結果テキスト/K01.txt')
# os.mkdir('GNP_result/'+'netview_message_search')
# rwfile('検索結果テキスト/K90GNP.txt','GNP_result/'+'netview_message',True)
def fileloop(text_path,output_path):
    my_dict=getname(dir_path) 
    if not path.exists(output_path+'/keyword_search_result'):
        mkdir(output_path+'/keyword_search_result')

    for file in listdir(text_path):

        if file.endswith('GNP.txt') or file.endswith('GNP.TXT'):
            if path.exists(output_path+'/keyword_search_result/GNP_result/'+file[1:3]+'_'+my_dict[file[1:3]]):
                print(file[1:3]+'_'+my_dict[file[1:3]]+'is existed')
            else:
                makedirs(output_path+'/keyword_search_result/GNP_result/'+file[1:3]+'_'+my_dict[file[1:3]])
            rwfile(text_path+'/'+file,output_path+'/keyword_search_result/GNP_result/'+file[1:3]+'_'+my_dict[file[1:3]],True)
        else:
            if path.exists(output_path+'/keyword_search_result/result/'+file[1:3]+'_'+my_dict[file[1:3]]):
                print(file[1:3]+'_'+my_dict[file[1:3]]+'is existed')
            else:   
                makedirs(output_path+'/keyword_search_result/result/'+file[1:3]+'_'+my_dict[file[1:3]])
            #print(file)
            rwfile(text_path+'/'+file,output_path+'/keyword_search_result/result/'+file[1:3]+'_'+my_dict[file[1:3]],False)

text_path='C:/vscode_workspace\项目组相关\検索結果テキスト1'
output_path='C:/vscode_workspace\项目组相关\项目组自动化'
#fileloop(text_path,output_path)
#fileloop('検索結果テキスト')


